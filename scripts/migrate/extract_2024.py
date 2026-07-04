import datetime
import openpyxl
from .parsing import read_numeric_row, is_numeric

FILE = "ARCHIVED - Milfore Invitational 2024.xlsx"


def _read_holes(ws, header_row, start_col, n_holes=18):
    pars = read_numeric_row(ws, header_row + 1, start_col, n_holes)
    yards = read_numeric_row(ws, header_row + 2, start_col, n_holes)
    sidx = read_numeric_row(ws, header_row + 3, start_col, n_holes)
    return [
        {"hole_no": i + 1, "par": pars[i], "yardage": yards[i], "stroke_index": sidx[i]}
        for i in range(n_holes)
    ]


def _embedded_leaderboard(ws, max_row=20):
    """Every 2024 round tab has its own 'Leaderboard:' block near the top with
    columns like Rank / Tie / 'Pair / Team' (or 'Player - Team') / Current Net /
    Tournament Points Awarded. Column offsets vary slightly per round, so locate
    them by header text instead of hardcoding positions.
    Returns list of {label, team_name, net_score, points, rank}."""
    header_row = None
    for r in range(1, max_row):
        for c in range(1, 30):
            v = ws.cell(row=r, column=c).value
            if v == "Tournament Points Awarded":
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        return []

    label_col = net_col = points_col = None
    for c in range(1, 30):
        v = ws.cell(row=header_row + 1, column=c).value
        if isinstance(v, str) and ("Team" in v or "Pair" in v or "Player" in v):
            label_col = c
    for r in (header_row, header_row + 1):
        for c in range(1, 30):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("Current Net"):
                net_col = c
            if v == "Tournament Points Awarded":
                points_col = c
    if not (label_col and net_col and points_col):
        return []

    entries = []
    r = header_row + 2
    while True:
        label = ws.cell(row=r, column=label_col).value
        if not isinstance(label, str) or " - Team " not in label:
            break
        net = ws.cell(row=r, column=net_col).value
        pts = ws.cell(row=r, column=points_col).value
        name_part, team_name = label.rsplit(" - ", 1)
        rank = r - header_row - 1
        entries.append({
            "label": name_part, "team_name": team_name,
            "event_score": net if is_numeric(net) else None,
            "points": pts if is_numeric(pts) else None, "rank": rank,
        })
        r += 1
    return entries


def _pair_rows(ws, row1, name_col, hdcp_col, score_col, n_holes=18):
    players = []
    for r in (row1, row1 + 1):
        name = ws.cell(row=r, column=name_col).value
        hdcp = ws.cell(row=r, column=hdcp_col).value
        scores = read_numeric_row(ws, r, score_col, n_holes)
        players.append({"name": name, "handicap_index": hdcp if is_numeric(hdcp) else None, "gross_scores": scores})
    return players


def extract():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    rounds = []

    # ---- Round 1: 2-Man Shamble (Cha Cha Cha, NASCAR), Sundance ----
    # Individual scoring (NOT shared): each partner has their own 75% handicap
    # allowance and own gross/net score. Column layout: B=name, C=raw HDCP,
    # scores start col K (11).
    ws = wb["1. Sundance - Shamble (NASCAR)"]
    holes = _read_holes(ws, 10, 11)
    lb = _embedded_leaderboard(ws)
    # Row spacing between pairs/teams is irregular (a team's 2 pairs sit close
    # together, then a gap before the next team's block). Self-locate every
    # player row instead of assuming fixed offsets: any row with a name in
    # col B and a numeric score in col K (11) is a player's own gross-score row.
    players_by_name = {}
    for r in range(17, 200):
        name = ws.cell(row=r, column=2).value
        first_score = ws.cell(row=r, column=11).value
        if not isinstance(name, str) or not is_numeric(first_score):
            continue
        hdcp = ws.cell(row=r, column=3).value
        scores = read_numeric_row(ws, r, 11, 18)
        players_by_name[name] = {"name": name, "handicap_index": hdcp if is_numeric(hdcp) else None, "gross_scores": scores}
    matches = []
    # group leaderboard entries into matches of 2 (pair label "A / B")
    match_number = 1.1
    seen = set()
    for entry in lb:
        if entry["label"] in seen:
            continue
        names = [n.strip() for n in entry["label"].split(" / ")]
        p_list = [players_by_name[n] for n in names if n in players_by_name]
        if not p_list:
            continue
        matches.append({
            "match_number": round(match_number, 1),
            "team_players": {entry["team_name"]: p_list},
            "leaderboard_entries": [entry],
        })
        match_number += 0.1
        seen.add(entry["label"])
    rounds.append({
        "round_number": 1, "course_name": "Sundance", "event_category": "2-Man",
        "event_name": "Cha Cha Cha (Shamble)", "scoring_type": "Stroke Play",
        "format_name": "2-Man Shamble (2024/25 - 75% Individual)", "handicap_adjusted": True,
        "round_date": datetime.date(2024, 8, 8), "tee_color": None, "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    # ---- Round 2: 2-Man Scramble, Antrim Dells ----
    ws = wb["2. Antrim Dells - Scramble (NAS"]
    holes = _read_holes(ws, 10, 11)
    lb = _embedded_leaderboard(ws)
    match_number = 2.1
    matches = []
    # Pairs occupy 2 CONSECUTIVE rows (row, row+1) sharing one team score, but
    # row spacing between pairs is irregular. The pair label ("A / B") only
    # appears in col J (10) on a pair's first row -- use that to find starts.
    pair_starts = []
    for r in range(17, 100):
        label = ws.cell(row=r, column=10).value
        if isinstance(label, str) and " / " in label:
            pair_starts.append(r)
    for r1 in pair_starts:
        p = _pair_rows(ws, r1, name_col=2, hdcp_col=3, score_col=11, n_holes=18)
        names = [p[0]["name"].strip(), p[1]["name"].strip()]
        entry = next((e for e in lb if set(n.strip() for n in e["label"].split(" / ")) == set(names)), None)
        team_name = entry["team_name"] if entry else None
        matches.append({
            "match_number": round(match_number, 1),
            "team_players": {team_name: p},
            "leaderboard_entries": [entry] if entry else [],
        })
        match_number += 0.1
    rounds.append({
        "round_number": 2, "course_name": "Antrim Dells", "event_category": "2-Man",
        "event_name": "Scramble", "scoring_type": "Stroke Play",
        "format_name": "2-Man Scramble", "handicap_adjusted": True,
        "round_date": datetime.date(2024, 8, 9), "tee_color": None, "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    # ---- Round 3: Team Stableford (95% handicap, 3 teams of 4), Charlevoix CC ----
    ws = wb["3. Charlevoix CC - Team Stablef"]
    holes = _read_holes(ws, 7, 11)
    team_rows = {"Team 1": [14, 17, 20, 23], "Team 2": [29, 32, 35, 38], "Team 3": [44, 47, 50, 53]}
    team_players = {}
    team_totals = {}
    for team_name, rows_ in team_rows.items():
        players = []
        for r in rows_:
            name = ws.cell(row=r, column=10).value
            hdcp_raw = None
            f_val = ws.cell(row=r, column=6).value  # F = hdcp*0.95 (pre-computed, but we recompute ourselves too)
            scores = read_numeric_row(ws, r, 11, 18)
            # back out raw handicap_index from the *0.95 pre-computed value when clean
            if is_numeric(f_val):
                hdcp_raw = round(f_val / 0.95, 2)
            players.append({"name": name, "handicap_index": hdcp_raw, "gross_scores": scores})
        team_players[team_name] = players
    # team stableford point totals from '0. Leaderboard' (rows 36,43,50 -- "Team N - Overall")
    lbws = wb["0. Leaderboard"]
    for r in range(1, lbws.max_row + 1):
        b = lbws.cell(row=r, column=5).value
        if isinstance(b, str) and b.endswith("- Overall"):
            team_name = lbws.cell(row=r, column=6).value
            total = lbws.cell(row=r, column=10).value
            if is_numeric(total):
                team_totals[team_name] = total
    ranked = sorted(team_totals.items(), key=lambda kv: -kv[1])
    points_by_team = {}
    rank_points = {1: 4, 2: 2, 3: 0}  # confirmed 2024 rank->points
    for i, (team_name, total) in enumerate(ranked):
        points_by_team[team_name] = {"points": rank_points.get(i + 1, 0), "rank": i + 1, "event_score": total}
    # "Top Individual Finisher" bonus (1 pt) -- INDIVIDUAL rank 1 (col C) across
    # the whole round, not a team-level award, so it needs its own player lookup.
    # Scope the search to the Round 3 section only (rank=1 recurs every round).
    round3_start = round3_end = None
    for r in range(1, lbws.max_row + 1):
        a = lbws.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("Round "):
            n = int(a.replace("Round ", "").strip().split()[0])
            if n == 3:
                round3_start = r
            elif round3_start and round3_end is None:
                round3_end = r
    round3_end = round3_end or (lbws.max_row + 1)
    top_individual = None
    for r in range(round3_start, round3_end):
        rank_adj = lbws.cell(row=r, column=3).value
        name = lbws.cell(row=r, column=5).value
        if rank_adj == 1 and isinstance(name, str) and not name.endswith("- Overall"):
            top_individual = name
            break
    rounds.append({
        "round_number": 3, "course_name": "Charlevoix CC", "event_category": "Individual",
        "event_name": "Team Stableford", "scoring_type": "Stableford",
        "format_name": "Team Stableford (2024/25 - 95% Handicap, 3-Team)", "handicap_adjusted": True,
        "round_date": datetime.date(2024, 8, 9), "tee_color": None, "holes": holes,
        "matches": [{
            "match_number": 3.1,
            "team_players": team_players,
            "leaderboard_entries": [
                {"label": t, "team_name": t, "event_score": points_by_team[t]["event_score"],
                 "points": points_by_team[t]["points"], "rank": points_by_team[t]["rank"]}
                for t in team_players
            ],
        }],
        "bonus_name": "Top Individual Finisher",
        "individual_bonus": {"player_name": top_individual, "points": 1} if top_individual else None,
    })

    # ---- Round 4: Solo Stroke Play (95% handicap), Torch Lake ----
    ws = wb["4. Torch Lake - Solo Stroke Pla"]
    holes = _read_holes(ws, 16, 10, n_holes=18)
    lb = _embedded_leaderboard(ws)
    matches = []
    match_number = 4.1
    for row in range(20, 100):
        name = ws.cell(row=row, column=9).value
        d_val = ws.cell(row=row, column=4).value  # handicap*0.95, pre-computed; also filters out 'Net' rows
        if not isinstance(name, str) or name in ("Player", "Net") or not is_numeric(d_val):
            continue
        hdcp_raw = round(d_val / 0.95, 2)
        scores = read_numeric_row(ws, row, 10, 18)
        entry = next((e for e in lb if e["label"].strip() == name.strip()), None)
        team_name = entry["team_name"] if entry else None
        matches.append({
            "match_number": round(match_number, 1),
            "team_players": {team_name: [{"name": name, "handicap_index": hdcp_raw, "gross_scores": scores}]},
            "leaderboard_entries": [entry] if entry else [],
        })
        match_number += 0.1
    rounds.append({
        "round_number": 4, "course_name": "Torch Lake", "event_category": "Individual",
        "event_name": "Head to Head (Stroke Play)", "scoring_type": "Stroke Play",
        "format_name": "Individual Stroke Play", "handicap_adjusted": True,
        "round_date": datetime.date(2024, 8, 10), "tee_color": None, "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    wb.close()
    return rounds
