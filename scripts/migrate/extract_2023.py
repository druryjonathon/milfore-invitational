import datetime
import openpyxl
from .parsing import read_numeric_row, is_numeric

FILE = "ARCHIVED - Milfore Invitational 2023.xlsx"


def _leaderboard_blocks(wb):
    """Parse '0. Leaderboard' into {round_number: {match_number: [entry, entry, ...]}}
    where entry = {label, team_name, event_score, rank, points, overall_rank, bonus_pts}.
    Entries within a match are in the SAME order they appear on the sheet, which is
    always Team 1 first then Team 2 -- matching the order used when reading each
    round tab's own Team 1 / Team 2 blocks."""
    ws = wb["0. Leaderboard"]
    result = {}
    current_round = None
    current_match = None
    for row in range(1, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        if isinstance(a, str) and a.startswith("Round "):
            current_round = int(a.replace("Round ", "").strip().split()[0])
            result[current_round] = {}
            current_match = None
            # fall through: this row may ALSO carry the first "Matchup X.Y" label in col B
        b = ws.cell(row=row, column=2).value
        if isinstance(b, str) and b.startswith("Matchup "):
            current_match = float(b.replace("Matchup ", "").strip())
            result[current_round][current_match] = []
            continue
        c = ws.cell(row=row, column=3).value  # team name
        d = ws.cell(row=row, column=4).value  # metric value (net strokes / stableford pts / holes won)
        if current_round is not None and current_match is not None and is_numeric(d):
            e = ws.cell(row=row, column=5).value
            g = ws.cell(row=row, column=7).value
            i = ws.cell(row=row, column=9).value
            j = ws.cell(row=row, column=10).value
            result[current_round][current_match].append({
                "label": b, "team_name": c, "event_score": d, "rank": e,
                "points": g, "overall_rank": i, "bonus_pts": j,
            })
    return result


def _read_holes(ws, header_row, start_col, n_holes=18):
    pars = read_numeric_row(ws, header_row + 1, start_col, n_holes)
    yards = read_numeric_row(ws, header_row + 2, start_col, n_holes)
    sidx = read_numeric_row(ws, header_row + 3, start_col, n_holes)
    return [
        {"hole_no": i + 1, "par": pars[i], "yardage": yards[i], "stroke_index": sidx[i]}
        for i in range(n_holes)
    ]


def _pair_rows(ws, row1, name_col, hdcp_col, score_col, n_holes=18):
    players = []
    for r in (row1, row1 + 1):
        name = ws.cell(row=r, column=name_col).value
        hdcp = ws.cell(row=r, column=hdcp_col).value
        scores = read_numeric_row(ws, r, score_col, n_holes)
        players.append({"name": name, "handicap_index": hdcp if is_numeric(hdcp) else None, "gross_scores": scores})
    return players


def _extract_2man_round(ws, round_number, course_name, event_name, first_matchup_row,
                         match_numbers, name_col=2, hdcp_col=3, score_col=9, header_row=7,
                         round_date=None, tee_color="Blue Tees", lb=None):
    holes = _read_holes(ws, header_row, score_col)
    matches = []
    lb_blocks = lb.get(round_number, {})
    for match_number, row1 in zip(match_numbers, range(first_matchup_row, first_matchup_row + 8 * len(match_numbers), 8)):
        team1 = _pair_rows(ws, row1, name_col, hdcp_col, score_col)
        team2 = _pair_rows(ws, row1 + 2, name_col, hdcp_col, score_col)
        entries = lb_blocks.get(match_number, [])
        matches.append({
            "match_number": match_number,
            "team_players": {"Team 1": team1, "Team 2": team2},
            "leaderboard_entries": entries,  # [team1_entry, team2_entry]
        })
    return {
        "round_number": round_number, "course_name": course_name, "event_category": "2-Man",
        "event_name": event_name, "scoring_type": "Stroke Play", "format_name": "2-Man Scramble",
        "handicap_adjusted": True, "round_date": round_date, "tee_color": tee_color,
        "holes": holes, "matches": matches,
    }


def extract():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    lb = _leaderboard_blocks(wb)
    rounds = []

    # ---- Round 1: 2-Man Shamble (Cha Cha Cha), Pine Knob ----
    ws = wb["1. Pine Knob - Shamble"]
    rounds.append(_extract_2man_round(
        ws, 1, "Pine Knob", "Cha Cha Cha (Shamble)", first_matchup_row=14,
        match_numbers=[1.1, 1.2, 1.3], round_date=datetime.date(2023, 8, 3), lb=lb,
    ))
    rounds[-1]["bonus_name"] = "Low Net"

    # ---- Round 2: 2-Man Scramble, Shepherds Hollow ----
    ws = wb["2. Shepherds Hollow - Scramble"]
    rounds.append(_extract_2man_round(
        ws, 2, "Shepherds Hollow", "Scramble", first_matchup_row=14,
        match_numbers=[2.1, 2.2, 2.3], round_date=datetime.date(2023, 8, 4), lb=lb,
    ))
    rounds[-1]["bonus_name"] = "Low Net"

    # ---- Round 3: Team Stableford (gross, no handicap), Shepherds Hollow ----
    ws = wb["3. Shepherds Hollow - Team Stab"]
    holes = _read_holes(ws, 7, 3)  # header at row7, holes start col C(3)
    team1_rows = [14, 16, 18, 20, 22, 24]
    team2_rows = [29, 31, 33, 35, 37, 39]

    def read_stableford_player(r):
        name = ws.cell(row=r, column=2).value
        scores = read_numeric_row(ws, r, 3, 18)
        return {"name": name, "handicap_index": None, "gross_scores": scores}

    team1 = [read_stableford_player(r) for r in team1_rows]
    team2 = [read_stableford_player(r) for r in team2_rows]
    entries = lb.get(3, {}).get(3.1, [])
    rounds.append({
        "round_number": 3, "course_name": "Shepherds Hollow", "event_category": "Individual",
        "event_name": "Team Stableford", "scoring_type": "Stableford",
        "format_name": "Team Stableford (2023 - Gross, 2-Team)", "handicap_adjusted": False,
        "round_date": datetime.date(2023, 8, 4), "tee_color": "Blue Tees", "holes": holes,
        "matches": [{
            "match_number": 3.1,
            "team_players": {"Team 1": team1, "Team 2": team2},
            "leaderboard_entries": entries,
        }],
    })

    # ---- Round 4: Match Play 1v1, Fieldstone ----
    ws = wb["4. Fieldstone - 1v1 Match Play"]
    holes = _read_holes(ws, 10, 6)  # header at row10, holes start col F(6)
    mp_rows = {4.1: (17, 19), 4.2: (28, 30), 4.3: (39, 41), 4.4: (50, 52), 4.5: (61, 63), 4.6: (72, 74)}
    matches = []
    for match_number, (r1, r2) in mp_rows.items():
        def read_mp_player(r):
            # column B is a RANK formula, C is handicap, E is the player name (confirmed via formula trace)
            name = ws.cell(row=r, column=5).value
            hdcp = ws.cell(row=r, column=3).value
            scores = read_numeric_row(ws, r, 6, 18)
            return {"name": name, "handicap_index": hdcp if is_numeric(hdcp) else None, "gross_scores": scores}
        p1 = read_mp_player(r1)
        p2 = read_mp_player(r2)
        entries = lb.get(4, {}).get(match_number, [])
        matches.append({
            "match_number": match_number,
            "team_players": {"Team 1": [p1], "Team 2": [p2]},
            "leaderboard_entries": entries,
        })
    rounds.append({
        "round_number": 4, "course_name": "Fieldstone", "event_category": "Individual",
        "event_name": "Head to Head (Match Play)", "scoring_type": "Match Play",
        "format_name": "Match Play 1v1 (2023)", "handicap_adjusted": True,
        "round_date": datetime.date(2023, 8, 5), "tee_color": "Gold Tees", "holes": holes,
        "matches": matches, "bonus_name": "Most Holes Won",
    })

    wb.close()
    return rounds
