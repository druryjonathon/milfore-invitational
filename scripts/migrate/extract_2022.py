import datetime
import openpyxl
from .parsing import read_numeric_row, is_numeric

FILE = "ARCHIVED - Milfore Invitational 2022.xlsx"

TEAM_NAMES = {"Team 1", "Team 2", "Team 3"}


def _read_holes(ws, header_row, start_col, n_holes):
    pars = read_numeric_row(ws, header_row + 1, start_col, n_holes)
    yards = read_numeric_row(ws, header_row + 2, start_col, n_holes)
    sidx = read_numeric_row(ws, header_row + 3, start_col, n_holes)
    return [
        {"hole_no": i + 1, "par": pars[i], "yardage": yards[i], "stroke_index": sidx[i]}
        for i in range(n_holes)
    ]


def _leaderboard_rounds(wb):
    """2022's 'Leaderboard' sheet: one row per match (pair/team/player), NOT
    grouped across multiple rows like later years. Returns {round_number: [entries]}
    where entry = {label, team_name, event_score(net strokes), rank, points}."""
    ws = wb["Leaderboard"]
    result = {}
    current_round = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("Round "):
            current_round = int(a.replace("Round ", "").strip().split()[0])
            result[current_round] = []
        team = ws.cell(row=r, column=3).value
        strokes = ws.cell(row=r, column=4).value
        if current_round is not None and team in TEAM_NAMES and is_numeric(strokes):
            label = ws.cell(row=r, column=2).value
            rank = ws.cell(row=r, column=5).value
            points = ws.cell(row=r, column=6).value
            result[current_round].append({
                "label": label, "team_name": team, "event_score": strokes,
                "rank": rank, "points": points,
            })
    return result


def _player_team(wb):
    """Static season-long roster + raw handicaps from the 'Teams' sheet."""
    ws = wb["Teams"]
    handicaps = {}
    for r in range(2, 14):
        name = ws.cell(row=r, column=2).value  # shorthand (matches round-tab labels)
        hdcp = ws.cell(row=r, column=3).value
        if isinstance(name, str) and is_numeric(hdcp):
            handicaps[name] = hdcp
    team_of = {}
    team_cols = {"Team 1": 6, "Team 2": 7, "Team 3": 8}  # F, G, H
    for team_name, col in team_cols.items():
        for r in range(4, 8):
            full_name = ws.cell(row=r, column=col).value
            if isinstance(full_name, str):
                # match against handicaps dict by shorthand suffix (last name / nickname)
                shorthand = next((n for n in handicaps if full_name.endswith(n) or n in full_name), None)
                if shorthand:
                    team_of[shorthand] = team_name
    return handicaps, team_of


def _match_entries_by_row(ws, rows, name_col, score_col, n_holes, hdcp_col=None):
    """Read one player/pair/group's scores starting at `rows[0]`; if hdcp_col is
    given, individual per-player rows are read directly (round 2/4 style).
    Otherwise the label itself may be a ' / '-joined pair/group sharing one
    scramble score (round 1/3 style)."""
    out = []
    for r in rows:
        name = ws.cell(row=r, column=name_col).value
        if not isinstance(name, str):
            continue
        hdcp = ws.cell(row=r, column=hdcp_col).value if hdcp_col else None
        scores = read_numeric_row(ws, r, score_col, n_holes)
        out.append({"name": name, "handicap_index": hdcp if is_numeric(hdcp) else None, "gross_scores": scores})
    return out


def extract():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    lb = _leaderboard_rounds(wb)
    handicaps, team_of = _player_team(wb)
    rounds = []

    # ---- Round 1: 2-Man Scramble ("1/4's and 2/3's"), Schuss Mountain ----
    ws = wb["1. Scorecard - Schuss Mountain"]
    holes = _read_holes(ws, 7, 3, 18)
    matches = []
    for i, r1 in enumerate([14, 16, 18, 24, 26, 28]):
        label = ws.cell(row=r1, column=2).value
        names = [n.strip() for n in label.split(" / ")]
        scores = read_numeric_row(ws, r1, 3, 18)
        players = [{"name": n, "handicap_index": handicaps.get(n), "gross_scores": scores} for n in names]
        entry = next((e for e in lb.get(1, []) if e["label"] == label), None)
        team_name = entry["team_name"] if entry else team_of.get(names[0])
        matches.append({
            "match_number": round(1.1 + i * 0.1, 1),
            "team_players": {team_name: players},
            "leaderboard_entries": [entry] if entry else [],
        })
    rounds.append({
        "round_number": 1, "course_name": "Schuss Mountain", "event_category": "2-Man",
        "event_name": "Scramble (1/4's & 2/3's)", "scoring_type": "Stroke Play",
        "format_name": "2-Man Scramble", "handicap_adjusted": True,
        "round_date": datetime.date(2022, 8, 18), "tee_color": "White Tees", "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    # ---- Round 2: 2-Man Scramble ("1/2's & 3/4's"), Summit ----
    ws = wb["2. Scorecard - Summit"]
    holes = _read_holes(ws, 7, 8, 18)
    matches = []
    pair_starts = []
    # Real pairing rows only -- a duplicate summary section further down the
    # sheet (rows 35+) re-lists the same pair labels and must be excluded.
    for r in range(14, 30):
        label = ws.cell(row=r, column=7).value
        if isinstance(label, str) and " / " in label:
            pair_starts.append(r)
    for i, r1 in enumerate(pair_starts):
        p = []
        for r in (r1, r1 + 1):
            name = ws.cell(row=r, column=2).value
            hdcp = ws.cell(row=r, column=3).value
            scores = read_numeric_row(ws, r, 8, 18)
            p.append({"name": name, "handicap_index": hdcp if is_numeric(hdcp) else handicaps.get(name), "gross_scores": scores})
        label = ws.cell(row=r1, column=7).value
        entry = next((e for e in lb.get(2, []) if e["label"] == label), None)
        team_name = entry["team_name"] if entry else team_of.get(p[0]["name"])
        matches.append({
            "match_number": round(2.1 + i * 0.1, 1),
            "team_players": {team_name: p},
            "leaderboard_entries": [entry] if entry else [],
        })
    rounds.append({
        "round_number": 2, "course_name": "Summit", "event_category": "2-Man",
        "event_name": "Scramble (1/2's & 3/4's)", "scoring_type": "Stroke Play",
        "format_name": "2-Man Scramble", "handicap_adjusted": True,
        "round_date": datetime.date(2022, 8, 19), "tee_color": "Blue Tees", "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    # ---- Round 3: 4-Man Scramble, The Legend ----
    ws = wb["3. Scorecard - The Legend "]
    holes = _read_holes(ws, 7, 3, 18)
    matches = []
    for i, r1 in enumerate([14, 16, 18]):
        label = ws.cell(row=r1, column=2).value
        names = [n.strip() for n in label.split(" / ")]
        scores = read_numeric_row(ws, r1, 3, 18)
        players = [{"name": n, "handicap_index": handicaps.get(n), "gross_scores": scores} for n in names]
        entry = next((e for e in lb.get(3, []) if e["label"] == label), None)
        team_name = entry["team_name"] if entry else team_of.get(names[0])
        matches.append({
            "match_number": round(3.1 + i * 0.1, 1),
            "team_players": {team_name: players},
            "leaderboard_entries": [entry] if entry else [],
        })
    rounds.append({
        "round_number": 3, "course_name": "The Legend", "event_category": "4-Man",
        "event_name": "Scramble", "scoring_type": "Stroke Play",
        "format_name": "4-Man Scramble", "handicap_adjusted": True,
        "round_date": datetime.date(2022, 8, 19), "tee_color": "Black Tees", "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    # ---- Round 4: Individual Head-to-Head (raw handicap, tiered field), Cedar River ----
    ws = wb["4. Scorecard - Cedar River"]
    holes = _read_holes(ws, 7, 6, 18)
    matches = []
    i = 0
    for r in range(14, 100):
        name = ws.cell(row=r, column=5).value
        hdcp = ws.cell(row=r, column=3).value
        if isinstance(name, str) and name not in ("Player",) and is_numeric(hdcp):
            scores = read_numeric_row(ws, r, 6, 18)
            entry = next((e for e in lb.get(4, []) if e["label"] == name), None)
            team_name = entry["team_name"] if entry else team_of.get(name)
            # Each "tier" (3 players, one per team) is ONE match so the
            # match-scoped strokes-received field correctly pools all 3
            # tier-mates together, matching how Match Play matches are built.
            tier_idx = i // 3
            if tier_idx >= len(matches):
                matches.append({
                    "match_number": round(4.1 + tier_idx * 0.1, 1),
                    "team_players": {},
                    "leaderboard_entries": [],
                })
            matches[tier_idx]["team_players"][team_name] = [
                {"name": name, "handicap_index": hdcp, "gross_scores": scores}
            ]
            if entry:
                matches[tier_idx]["leaderboard_entries"].append(entry)
            i += 1
    rounds.append({
        "round_number": 4, "course_name": "Cedar River", "event_category": "Individual",
        "event_name": "Head to Head", "scoring_type": "Stroke Play",
        "format_name": "Individual Stroke Play (2022 - Raw Handicap, Tiered Field)", "handicap_adjusted": True,
        "round_date": datetime.date(2022, 8, 20), "tee_color": "Blue Tees", "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    wb.close()
    return rounds
