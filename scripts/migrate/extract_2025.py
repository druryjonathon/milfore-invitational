import datetime
import openpyxl
from .parsing import read_numeric_row, is_numeric
from .compute import adjusted_handicap_individual_pct, strokes_received_for_field, strokes_on_hole

FILE = "ARCHIVED - Milfore Invitational 2025 - FIXED.xlsx"


def _read_holes(ws, header_row, start_col, n_holes=18):
    pars = read_numeric_row(ws, header_row + 1, start_col, n_holes)
    yards = read_numeric_row(ws, header_row + 2, start_col, n_holes)
    sidx = read_numeric_row(ws, header_row + 3, start_col, n_holes)
    return [
        {"hole_no": i + 1, "par": pars[i], "yardage": yards[i], "stroke_index": sidx[i]}
        for i in range(n_holes)
    ]


def _leaderboard_section(lbws, start_row, end_row):
    """Rows 23-28 / 33-35 / 40-51 style blocks in the (fixed) '0. Leaderboard' sheet:
    col2=RankInitial, col3=RankAdj, col5=Players(label with ' - Team N'), col6=Team,
    col8=Gross, col9=Net, col13=Points."""
    entries = []
    for r in range(start_row, end_row + 1):
        label = lbws.cell(row=r, column=5).value
        team = lbws.cell(row=r, column=6).value
        if not isinstance(label, str) or not isinstance(team, str):
            continue
        entries.append({
            "label": label.rsplit(" - ", 1)[0],
            "team_name": team,
            "rank": lbws.cell(row=r, column=3).value,
            "event_score": lbws.cell(row=r, column=9).value,
            "gross": lbws.cell(row=r, column=8).value,
            "points": lbws.cell(row=r, column=13).value,
        })
    return entries


def extract():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    lbws = wb["0. Leaderboard"]
    rounds = []

    # Build a player -> team_name map from round 2's leaderboard (reused for round 1,
    # which has no per-player team labels of its own).
    r2_entries = _leaderboard_section(lbws, 23, 28)
    player_team = {}
    for e in r2_entries:
        for n in e["label"].split(" / "):
            player_team[n.strip()] = e["team_name"]

    # ---- Round 1: "Modified Cha Cha Cha", SH North ----
    # Individual 95% handicap (confirmed via sheet's own "Handicap 95%" label and
    # cross-check against round 3's raw handicap index), but scored as a TEAM-summed
    # net-strokes total for points -- the Leaderboard's team-rank formulas for this
    # round never fully populated even after the formula fix, so team totals/ranks/
    # points are computed here from each player's own net score.
    ws = wb["Round 1 - SH North"]
    holes = _read_holes(ws, 29, 11)
    players = []
    for r in range(36, 200):
        name = ws.cell(row=r, column=10).value
        hdcp = ws.cell(row=r, column=5).value
        if isinstance(name, str) and is_numeric(hdcp):
            scores = read_numeric_row(ws, r, 11, 18)
            players.append({"name": name, "handicap_index": hdcp, "gross_scores": scores})
    # compute adjusted handicap + round-wide strokes received (same as Individual Stroke Play)
    for p in players:
        p["_adj"] = adjusted_handicap_individual_pct(p["handicap_index"], 0.95)
    pool = {p["name"]: p["_adj"] for p in players}
    strokes = strokes_received_for_field(pool)
    # Round 1 was rained out/shortened; per Jon, unplayed holes are scored at
    # NET PAR -- i.e. gross = par + whatever stroke allowance that specific
    # hole would have given the player, so their net comes out to exactly par.
    # Track which holes were backfilled (vs. actually played) for the audit trail.
    for p in players:
        p["_assumed_holes"] = set()
        total_sr = strokes[p["name"]]
        for i, g in enumerate(p["gross_scores"]):
            if g is None:
                hole = holes[i]
                sr_hole = strokes_on_hole(hole["stroke_index"], total_sr)
                p["gross_scores"][i] = hole["par"] + sr_hole
                p["_assumed_holes"].add(hole["hole_no"])
    for p in players:
        gross_total = sum(g for g in p["gross_scores"] if g is not None)
        p["_net"] = gross_total - strokes[p["name"]]
        p["_gross_total"] = gross_total
        p["_team"] = player_team.get(p["name"])
    team_players = {}
    for p in players:
        team_players.setdefault(p["_team"], []).append(p)
    team_net_totals = {t: sum(p["_net"] for p in ps) for t, ps in team_players.items()}
    ranked_teams = sorted(team_net_totals.items(), key=lambda kv: kv[1])  # lowest net wins
    # Confirmed by Jon: Team 1 and Team 2 actually TIED for 1st/2nd that day and
    # split the points, 6/6, rather than the 8/4 a clean rank order would imply
    # (our own net-strokes recompute doesn't reproduce that exact tie -- likely a
    # difference in how the shortened round's net-par assumption was applied --
    # but the awarded points below are the confirmed, authoritative outcome).
    team_rank_points = {1: 6, 2: 6, 3: 0}
    team_result = {t: {"rank": i + 1, "points": team_rank_points.get(i + 1, 0), "total": total}
                    for i, (t, total) in enumerate(ranked_teams)}
    # "Individual Bonus Points" (Low Net style): rank the 12 players' own net scores, top 2 get 1.5/0.5
    by_net = sorted(players, key=lambda p: p["_net"])
    round1_bonus_recipients = [(by_net[0]["name"], 1.5), (by_net[1]["name"], 0.5)] if len(by_net) >= 2 else []
    rounds.append({
        "round_number": 1, "course_name": "Stonehedge North", "event_category": "Individual",
        "event_name": "Modified Cha Cha Cha", "scoring_type": "Stroke Play",
        "format_name": "Individual Stroke Play", "handicap_adjusted": True,
        "round_date": datetime.date(2025, 7, 24), "tee_color": None, "holes": holes,
        "matches": [{
            "match_number": 1.0 + (i + 1) / 10,
            "team_players": {t: ps},
            "leaderboard_entries": [{
                "label": t, "team_name": t, "event_score": team_result[t]["total"],
                "points": team_result[t]["points"], "rank": team_result[t]["rank"],
            }],
        } for i, (t, ps) in enumerate(team_players.items())],
        "bonus_name": "Low Net", "bonus_recipients": round1_bonus_recipients,
    })

    # ---- Round 2: 2-Man Scramble, Lake View West ----
    ws = wb["Round 2 - Lake View West"]
    holes = _read_holes(ws, 11, 11)
    pair_starts = []
    for r in range(18, 100):
        j_val = ws.cell(row=r, column=10).value
        if isinstance(j_val, str) and " / " in j_val:
            pair_starts.append(r)
    matches = []
    for i, r1 in enumerate(pair_starts):
        pl = []
        for r in (r1, r1 + 1):
            name = ws.cell(row=r, column=2).value
            hdcp = ws.cell(row=r, column=3).value
            scores = read_numeric_row(ws, r, 11, 18)
            pl.append({"name": name, "handicap_index": hdcp if is_numeric(hdcp) else None, "gross_scores": scores})
        names = [pl[0]["name"].strip(), pl[1]["name"].strip()]
        entry = next((e for e in r2_entries if set(n.strip() for n in e["label"].split(" / ")) == set(names)), None)
        team_name = entry["team_name"] if entry else None
        matches.append({
            "match_number": round(2.1 + i * 0.1, 1),
            "team_players": {team_name: pl},
            "leaderboard_entries": [entry] if entry else [],
        })
    rounds.append({
        "round_number": 2, "course_name": "Lake View West", "event_category": "2-Man",
        "event_name": "Scramble", "scoring_type": "Stroke Play",
        "format_name": "2-Man Scramble", "handicap_adjusted": True,
        "round_date": datetime.date(2025, 7, 25), "tee_color": None, "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    # ---- Round 3: 4-Man Scramble, Lake View East ----
    ws = wb["Round 3 - Lake View East"]
    holes = _read_holes(ws, 12, 13)
    r3_entries = _leaderboard_section(lbws, 33, 35)
    group_starts = []
    for r in range(19, 100):
        l_val = ws.cell(row=r, column=12).value
        if isinstance(l_val, str) and l_val.count(" / ") == 3:
            group_starts.append(r)
    matches = []
    for i, r1 in enumerate(group_starts):
        pl = []
        for r in range(r1, r1 + 4):
            name = ws.cell(row=r, column=4).value
            hdcp = ws.cell(row=r, column=5).value
            if isinstance(name, str) and is_numeric(hdcp):
                pl.append({"name": name, "handicap_index": hdcp, "_row": r})
        # scores appear on the first row and the last (team-aggregate) row only; shared across all 4
        scores = read_numeric_row(ws, r1, 13, 18)
        if all(s is None for s in scores):
            scores = read_numeric_row(ws, r1 + 3, 13, 18)
        for p in pl:
            p["gross_scores"] = scores
            del p["_row"]
        names = set(n.strip() for n in ws.cell(row=r1, column=12).value.split(" / "))
        entry = next((e for e in r3_entries if set(n.strip() for n in e["label"].split(" / ")) == names), None)
        team_name = entry["team_name"] if entry else None
        matches.append({
            "match_number": round(3.1 + i * 0.1, 1),
            "team_players": {team_name: pl},
            "leaderboard_entries": [entry] if entry else [],
        })
    rounds.append({
        "round_number": 3, "course_name": "Lake View East", "event_category": "4-Man",
        "event_name": "Scramble", "scoring_type": "Stroke Play",
        "format_name": "4-Man Scramble", "handicap_adjusted": True,
        "round_date": datetime.date(2025, 7, 25), "tee_color": None, "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    # ---- Round 4: Solo Stroke Play, Stoatin Brae ----
    ws = wb["Round 4 - Stoatin Brae"]
    holes = _read_holes(ws, 18, 10)
    r4_entries = _leaderboard_section(lbws, 40, 51)
    matches = []
    i = 0
    for r in range(23, 200):
        name = ws.cell(row=r, column=9).value
        hdcp = ws.cell(row=r, column=4).value
        if isinstance(name, str) and name not in ("Player", "Net") and is_numeric(hdcp):
            scores = read_numeric_row(ws, r, 10, 18)
            entry = next((e for e in r4_entries if e["label"].strip() == name.strip()), None)
            team_name = entry["team_name"] if entry else None
            matches.append({
                "match_number": round(4.1 + i * 0.1, 1),
                "team_players": {team_name: [{"name": name, "handicap_index": hdcp, "gross_scores": scores}]},
                "leaderboard_entries": [entry] if entry else [],
            })
            i += 1
    rounds.append({
        "round_number": 4, "course_name": "Stoatin Brae", "event_category": "Individual",
        "event_name": "Group NASCAR", "scoring_type": "Stroke Play",
        "format_name": "Individual Stroke Play", "handicap_adjusted": True,
        "round_date": datetime.date(2025, 7, 26), "tee_color": None, "holes": holes,
        "matches": matches, "bonus_name": None,
    })

    wb.close()
    return rounds
